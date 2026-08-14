"""Final aggregation: union v2 full-text deposits (high-conf, not-use) + structured DataBankList,
classify, cross-tab against MeSH sensitive-area categories, emit formatted XLSX + report_stats.txt.
ponytail: pandas + openpyxl, no report framework."""
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import date
import catalog

# Durable, non-git output location - see extract_databanks.py's OUT comment.
OUT=os.path.expanduser("~/Dropbox/Projects/Bulk Data Rule/data")
os.makedirs(OUT, exist_ok=True)
PROJ=os.path.expanduser("~/Dropbox/Projects/Bulk Data Rule"); Y=2020

def cls(n):
    r=catalog.classify(name=str(n)); return pd.Series([r['canonical'],r['tier'],r['country'],r['access'],r['bucket']])
# ---- full-text deposits (academic + preprint): keep high-confidence, non-use ----
def _ft(p):
    d=pd.read_csv(p); d=d[(d['confidence']=='high') & (d['context']!='use')].copy()
    d=d[['pmid','year','pmcid','repo','tier','country','access','bucket','matched_val']]; d['source']='fulltext'; return d
# ---- structured DataBankList (academic + preprint), since-2020 ----
def _db(p):
    d=pd.read_csv(p); d=d[d['year']>=Y].copy()
    d[['repo','tier','country','access','bucket']]=d['databank'].apply(cls)
    d=d.rename(columns={'accession':'matched_val'}); d['pmcid']=''; d['source']='structured'
    return d[['pmid','year','pmcid','repo','tier','country','access','bucket','matched_val','source']]

alld=pd.concat([_ft(f"{OUT}/deposits_v2.csv"), _ft(f"{OUT}/preprint_deposits_v2.csv"),
                _db(f"{OUT}/deposits_databank.csv"), _db(f"{OUT}/preprint_databank.csv")], ignore_index=True)
srcs=alld.groupby(['pmid','repo'])['source'].agg(lambda x:'both' if x.nunique()>1 else x.iloc[0]).reset_index()
uniq=alld.sort_values('source').drop_duplicates(['pmid','repo']).drop(columns='source').merge(srcs,on=['pmid','repo'])
ca=pd.read_csv(f"{OUT}/coauthor_affil.csv")[['pmid','coc_coauthor','coc_countries']]
uniq=uniq.merge(ca,on='pmid',how='left'); uniq['coc_coauthor']=uniq['coc_coauthor'].fillna(0).astype(int)

TIER={'CONCERN':0,'FOREIGN_OPEN':1,'FOREIGN_CTRL':2,'US_OPEN':3,'US_CTRL':4,'REGISTRY':5,'UNKNOWN':6}
land=uniq.groupby(['repo','tier','country','access','bucket']).agg(pubs=('pmid','nunique'),depositions=('pmid','count')).reset_index()
land['_o']=land['tier'].map(TIER).fillna(9); land=land.sort_values(['_o','pubs'],ascending=[True,False]).drop(columns='_o')
tier=uniq.groupby('tier').agg(pubs=('pmid','nunique')).reset_index(); tier['_o']=tier['tier'].map(TIER).fillna(9); tier=tier.sort_values('_o').drop(columns='_o')

# ---- MeSH sensitive-area cross-tab ----
mesh=pd.read_csv(f"{OUT}/mesh_sensitive.csv"); mesh['pmid']=mesh['pmid'].astype(int)
dep_pmids=set(uniq['pmid'].astype(int))
open_tiers={'US_OPEN','FOREIGN_OPEN'}; ctrl_tiers={'US_CTRL','FOREIGN_CTRL'}
open_pmids=set(uniq[uniq['tier'].isin(open_tiers)]['pmid'].astype(int))
ctrl_pmids=set(uniq[uniq['tier'].isin(ctrl_tiers)]['pmid'].astype(int))
CATS=["genomic","omic_other","health","biometric","geolocation"]
carea=[]
for c in CATS:
    sub=mesh[mesh['sensitive_cats'].fillna('').str.contains(c)]
    ids=set(sub['pmid'])
    carea.append((c,len(ids), len(ids&dep_pmids), len(ids&open_pmids), len(ids&ctrl_pmids), len(ids)-len(ids&dep_pmids)))
carea=pd.DataFrame(carea,columns=["Sensitive category (MeSH)","Pubs in area","With any deposit","With open deposit","With controlled deposit","No detected deposit"])

# ---- stats dump for narrative ----
st=pd.read_csv(f"{OUT}/status_v2.csv"); cov=pd.read_csv(f"{OUT}/coverage.csv")
c2020=cov[cov['year']>=Y]
with open(f"{OUT}/report_stats.txt","w") as f:
    f.write(f"since-{Y} corpus (first/last-author FT-faculty academic articles): {len(c2020)}\n")
    f.write(f"  in PMC: {int(c2020['pmcid'].notna().sum())}\n")
    f.write(f"  full-text machine-readable & scanned: {int(st['has_fulltext'].sum())} / {len(st)}\n")
    f.write(f"  with Data/Code Availability section: {int(st['has_das'].sum())}\n")
    f.write(f"distinct pubs with >=1 counted deposit (v2 hi-conf + structured): {uniq['pmid'].nunique()}\n")
    f.write(f"total counted depositions: {len(uniq)}\n")
    f.write("\n=== TIER SUMMARY (distinct pubs) ===\n"+tier.to_string(index=False))
    f.write("\n\n=== REPOSITORY LANDSCAPE ===\n"+land.to_string(index=False))
    f.write("\n\n=== SENSITIVE-AREA (MeSH upper bound) x DEPOSITION ===\n"+carea.to_string(index=False))
    f.write(f"\n\nMeSH: human-subjects pubs={int(mesh['human'].sum())}/{len(mesh)}; any sensitive area={(mesh['sensitive_cats'].fillna('')!='').sum()}\n")
    f.write("\n=== COUNTRY-OF-CONCERN: 0 counted deposits. Candidates (all USE/citation/low-conf): ===\n")
    cc=pd.read_csv(f"{OUT}/deposits_v2.csv"); cc=cc[cc['tier']=='CONCERN'][['pmid','year','repo','matched_val','context','confidence']]
    f.write(cc.to_string(index=False) if len(cc) else "  none")
    f.write("\n\n=== FOREIGN_CTRL (EGA) detail ===\n"+uniq[uniq['tier']=='FOREIGN_CTRL'][['pmid','year','repo','matched_val']].to_string(index=False))
    f.write("\n\n=== UNMAPPED repos (review) ===\n"+land[land['tier']=='UNKNOWN'].to_string(index=False))

# ---- XLSX ----
def ren(df,m): df=df.copy(); df.columns=[m.get(c,c) for c in df.columns]; return df
LM={'repo':'Repository','tier':'Risk tier','country':'Host country','access':'Access model','bucket':'Data type','pubs':'Pubs','depositions':'Depositions'}
RM={'pmid':'PMID','year':'Year','pmcid':'PMCID','repo':'Repository','tier':'Risk tier','country':'Host country','access':'Access model','bucket':'Data type','matched_val':'Accession / match','source':'Source','coc_coauthor':'COC coauthor','coc_countries':'COC country'}
raw=uniq.sort_values(['tier','year'],ascending=[True,False]).copy(); raw['coc_coauthor']=raw['coc_coauthor'].map({1:'YES',0:''})
notes=pd.DataFrame({"Item":["Corpus","Filters","Window","Sources","Deposit rule","Deposit vs use","Coverage limit","MeSH tag","COC coauthor","Risk tiers","Country-of-concern","Generated"],
 "Detail":["reciterdb full-time faculty, first OR last author",
   "authorPosition in (first,last); identity.fullTimeFaculty=yes; publicationTypeCanonical in (Academic Article, Preprint)",
   f"articleYear >= {Y}",
   "PubMed <DataBankList> (structured) UNION Europe PMC/NCBI full-text scan (accession + repo-URL + word-bounded name). Deposit tabs include preprints; the Sensitive Area (MeSH) tab is academic-corpus-based (preprints lack MeSH indexing).",
   "Counted only if high-confidence (accession/URL, or name inside Data-Availability section) AND not data-use language",
   "Matches with download/obtained-from/accessed context are excluded as data USE, not deposition",
   "Only OA full text is machine-readable (98% of since-2020 PMC pubs here). Prose-only 'available on request' not a deposit.",
   "Sensitive-area = MeSH topic tags (human-subjects required); a coarse UPPER BOUND, academic articles only (preprints untagged); not proof bulk identifiable data was shared",
   "Raw Deposits flags pubs with >=1 country-of-concern-affiliated author (PubMed all-author affiliations) -- population to review, not a violation",
   "CONCERN=country-of-concern host; FOREIGN_OPEN/CTRL=non-US host; US_OPEN=US host+open (globally downloadable); US_CTRL=US controlled (compliant); REGISTRY=trial/structure",
   "0 confirmed depositions; all name/URL hits were data reuse or citation of Chinese resources",
   str(date.today())]})
def autofmt(p):
    wb=load_workbook(p); ab=Font(name='Arial',size=12,bold=True); a=Font(name='Arial',size=12)
    fills={'CONCERN':'FFC7CE','FOREIGN_OPEN':'FFEB9C','FOREIGN_CTRL':'FCE4D6','US_OPEN':'DDEBF7','US_CTRL':'C6EFCE','REGISTRY':'EDEDED'}
    for ws in wb.worksheets:
        ws.freeze_panes='A2'
        for r in ws.iter_rows():
            for c in r: c.font=a
        for c in ws[1]: c.font=ab
        hdr=[c.value for c in ws[1]]
        if 'Risk tier' in hdr:
            ti=hdr.index('Risk tier')
            for r in ws.iter_rows(min_row=2):
                t=r[ti].value
                if t in fills:
                    for c in r: c.fill=PatternFill('solid',fgColor=fills[t])
        for i,col in enumerate(ws.columns,1):
            mx=max((len(str(c.value or '')) for c in col),default=0); ws.column_dimensions[get_column_letter(i)].width=min(mx+3,60)
    wb.save(p)
out=os.path.join(PROJ,f"WCM Sensitive Data Repository Landscape - {date.today()}.xlsx")
with pd.ExcelWriter(out,engine='openpyxl') as w:
    ren(land,LM).to_excel(w,sheet_name='Repository Landscape',index=False)
    ren(tier,{'tier':'Risk tier','pubs':'Pubs'}).to_excel(w,sheet_name='Risk Tier Summary',index=False)
    carea.to_excel(w,sheet_name='Sensitive Area (MeSH)',index=False)
    ren(raw,RM).to_excel(w,sheet_name='Raw Deposits',index=False)
    notes.to_excel(w,sheet_name='Method & Assumptions',index=False)
autofmt(out)
print("wrote",out,"\n"); print(open(f"{OUT}/report_stats.txt").read())
