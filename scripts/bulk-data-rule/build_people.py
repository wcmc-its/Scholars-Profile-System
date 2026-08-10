"""Person- and article-level XLSX: who deposited what, where, with full metadata + granular data type.
ponytail: read the attributed CSV, merge accessions back, groupby for summaries, openpyxl format."""
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import date
import catalog

OUT=os.path.dirname(os.path.abspath(__file__)); PROJ=os.path.expanduser("~/Dropbox/Projects/Bulk Data Rule")
full=pd.read_csv(f"{OUT}/attributed_deposits.csv").rename(columns={'title_x':'title','title_y':'article_title'})

# ---- merge accessions back (pmid,repo -> joined accessions) ----
acc=[]
for p,valcol in [("deposits_v2.csv","matched_val"),("preprint_deposits_v2.csv","matched_val")]:
    d=pd.read_csv(f"{OUT}/{p}"); d=d[(d['confidence']=='high')&(d['context']!='use')]
    acc.append(d[['pmid','repo',valcol]].rename(columns={valcol:'accession'}))
for p in ["deposits_databank.csv","preprint_databank.csv"]:
    d=pd.read_csv(f"{OUT}/{p}"); d=d[d['year']>=2020] if 'year' in d else d
    d['repo']=d['databank'].apply(lambda n: catalog.classify(name=str(n))['canonical'])
    acc.append(d[['pmid','repo','accession']])
accdf=pd.concat(acc,ignore_index=True).dropna(subset=['accession'])
accdf=accdf.groupby(['pmid','repo'])['accession'].agg(lambda s:"; ".join(sorted(set(map(str,s)))[:6])).reset_index()
full=full.merge(accdf,on=['pmid','repo'],how='left')

# ---- coauthor country-of-concern affiliation vector (covered person on the paper) ----
ca=pd.read_csv(f"{OUT}/coauthor_affil.csv")[['pmid','coc_coauthor','coc_countries','coc_examples']]
full=full.merge(ca,on='pmid',how='left')
full['coc_coauthor']=full['coc_coauthor'].fillna(0).astype(int)
full['coc_flag']=full['coc_coauthor'].map({1:'YES',0:''})

# ---- non-exempt exposure: open x genomic/'omic x NO US-federal funding (categorical prohibition, exemption unlikely) ----
OPEN={'US_OPEN','FOREIGN_OPEN'}; OMIC={'genomic','omic_other'}
def nonexempt(r):
    if r['tier'] not in OPEN: return ''
    sens=set(str(r['sensitive_cats']).split('|')) & OMIC
    if not sens or bool(r['usgov_funded']): return ''
    return "open "+",".join(sorted(sens))+", no US-federal funding"
full['nonexempt_exposure']=full.apply(nonexempt,axis=1)
full['_ne']=full['nonexempt_exposure']!=''
full['_coc_openomic']=(full['coc_coauthor']==1) & full['tier'].isin(OPEN) & full['sensitive_cats'].fillna('').str.contains('genomic|omic_other')

# ---- detail tab ----
det=full.copy()
DET=['cwid','lastName','firstName','dept','division','title','position','repo','repo_detail','tier','country','access',
     'bucket','sensitive_cats','sensitive_subtypes','concern','nonexempt_exposure','coc_flag','coc_countries','coc_examples','nih_funded','usgov_funded','funders',
     'journal','impact','year','pubdate','pubtype','citNIH','citScopus','pmid','pmcid','doi','accession','article_title']
det=det[DET].sort_values(['tier','lastName','year'],ascending=[True,True,False])
DNAMES={'cwid':'CWID','lastName':'Last name','firstName':'First name','dept':'Department','division':'Division','title':'Title',
 'position':'Author position','repo':'Repository','repo_detail':'Repository detail','tier':'Risk tier','country':'Host country',
 'access':'Access model','bucket':'Data type','sensitive_cats':'Sensitive category','sensitive_subtypes':'Sensitive subtypes',
 'concern':'Concern reason','nonexempt_exposure':'Non-exempt exposure','coc_flag':'COC coauthor','coc_countries':'COC country','coc_examples':'COC affiliation (example)','nih_funded':'NIH funded','usgov_funded':'US-gov funded','funders':'Funders',
 'journal':'Journal','impact':'Impact factor','year':'Year','pubdate':'Pub date','pubtype':'Pub type',
 'citNIH':'Citations (NIH)','citScopus':'Citations (Scopus)','pmid':'PMID','pmcid':'PMCID','doi':'DOI','accession':'Accession(s)','article_title':'Article title'}

# ---- faculty summary ----
def agg(g):
    return pd.Series({
      'Last name':g['lastName'].iloc[0],'First name':g['firstName'].iloc[0],
      'Department':g['dept'].iloc[0],'Division':g['division'].iloc[0],'Title':g['title'].iloc[0],
      'Deposit pubs':g['pmid'].nunique(),'Total deposits':len(g),
      'Non-exempt open-omic':g['_ne'].sum(),
      'Pubs w/ COC coauthor':g[g['coc_coauthor']==1]['pmid'].nunique(),
      'COC coauthor + open-omic':g[g['_coc_openomic']]['pmid'].nunique(),
      'Concerning deposits':(g['concern'].fillna('')!='').sum(),
      'Foreign-hosted':g['tier'].isin(['FOREIGN_OPEN','FOREIGN_CTRL']).sum(),
      'Open (US+foreign)':g['tier'].isin(['US_OPEN','FOREIGN_OPEN']).sum(),
      'Controlled (compliant)':g['tier'].isin(['US_CTRL','FOREIGN_CTRL']).sum(),
      'Country-of-concern':(g['tier']=='CONCERN').sum(),
      'NIH-funded pubs':g.drop_duplicates('pmid')['nih_funded'].sum(),
      'Repositories':"; ".join(sorted(set(g['repo'].dropna()))),
      'Top data subtypes':"; ".join(sorted({x for s in g['sensitive_subtypes'].dropna() for x in s.split('|') if x})[:6]),
    })
fac=full[full['cwid'].notna()].groupby('cwid').apply(agg).reset_index()
fac=fac.rename(columns={'cwid':'CWID'}).sort_values(['Non-exempt open-omic','Concerning deposits'],ascending=False)

# ---- by department ----
dept=full[full['cwid'].notna()].groupby('dept').agg(**{
  'Faculty':('cwid','nunique'),'Deposit pubs':('pmid','nunique'),'Total deposits':('repo','count'),
  'Concerning':('concern',lambda s:(s.fillna('')!='').sum()),
  'Foreign-hosted':('tier',lambda s:s.isin(['FOREIGN_OPEN','FOREIGN_CTRL']).sum()),
  'Controlled':('tier',lambda s:s.isin(['US_CTRL','FOREIGN_CTRL']).sum())}).reset_index().rename(columns={'dept':'Department'})
dept=dept.sort_values('Concerning',ascending=False)

# ---- concerning-only + non-exempt exposure + country-of-concern coauthors ----
conc=det[det['concern'].fillna('')!='']
ne=det[det['nonexempt_exposure'].fillna('')!='']
coc_tab=det[det['coc_flag']=='YES'].sort_values(['tier','coc_countries','lastName'])

notes=pd.DataFrame({"Item":["Corpus","Person attribution","'Non-exempt exposure' (priority tab)","Country-of-concern coauthors","'Concerning' definition","Granularity","Funding","Metadata source","Limitations","Generated"],
 "Detail":["reciterdb full-time faculty, first/last author, Academic Article + Preprint, year>=2020",
   "Deposit attributed to each WCM full-time first/last author on the paper (PI-level accountability)",
   "Open deposition (US or foreign) of human genomic/'omic data with NO detected US-federal funding. This is the sharp edge: 'omic data carries a categorical prohibition the rule cannot cure, and absent a federal award the exemption is unlikely. Confirm each against the actual award terms.",
   "'COC coauthor' = >=1 author on the paper affiliated at an institution in a country of concern (from PubMed all-author affiliations; country token). This is the 'covered person with access to the dataset' vector, distinct from repository host country. A COC-affiliated coauthor is not itself a violation determination -- it is the population where the covered-person-access question arises.",
   "Country-of-concern host, OR foreign-hosted, OR open deposition of a sensitive data type (genomic/omic/health/biometric/geolocation)",
   "Data sub-types from MeSH (e.g. omic->transcriptomic/single-cell/epigenomic/proteomic/metabolomic); repos resolved (SRA human vs non-human, GEO expression/single-cell/methylation)",
   "NIH/US-gov funding from PubMed GrantList; NIH-authorized deposition may be EXEMPT under the rule",
   "DOI/journal/impact/citations from reciterdb; grants/abstract/MeSH from PubMed efetch",
   "Deposition-driven; MeSH sub-types are a topic-level signal, not proof of data content; US-person counts not observable",
   str(date.today())]})

def autofmt(p):
    wb=load_workbook(p); ab=Font(name='Arial',size=12,bold=True); a=Font(name='Arial',size=12)
    fills={'CONCERN':'FFC7CE','FOREIGN_OPEN':'FFEB9C','FOREIGN_CTRL':'FCE4D6','US_OPEN':'DDEBF7','US_CTRL':'C6EFCE','REGISTRY':'EDEDED'}
    for ws in wb.worksheets:
        ws.freeze_panes='A2'
        for r in ws.iter_rows():
            for c in r: c.font=a
        for c in ws[1]: c.font=ab
        hdr=[c.value for c in ws[1]]
        if 'Risk tier' in hdr:
            ti=hdr.index('Risk tier')
            for r in ws.iter_rows(min_row=2):
                if r[ti].value in fills:
                    for c in r: c.fill=PatternFill('solid',fgColor=fills[r[ti].value])
        for i,col in enumerate(ws.columns,1):
            mx=max((len(str(c.value or '')) for c in col),default=0); ws.column_dimensions[get_column_letter(i)].width=min(mx+3,55)
    wb.save(p)

out=os.path.join(PROJ,f"WCM Sensitive Data - People & Metadata - {date.today()}.xlsx")
with pd.ExcelWriter(out,engine='openpyxl') as w:
    ne.rename(columns=DNAMES).to_excel(w,sheet_name='Non-Exempt Exposure',index=False)
    coc_tab.rename(columns=DNAMES).to_excel(w,sheet_name='COC Coauthors',index=False)
    fac.to_excel(w,sheet_name='Faculty Summary',index=False)
    conc.rename(columns=DNAMES).to_excel(w,sheet_name='Concerning Deposits',index=False)
    det.rename(columns=DNAMES).to_excel(w,sheet_name='All Deposits (detail)',index=False)
    dept.to_excel(w,sheet_name='By Department',index=False)
    notes.to_excel(w,sheet_name='Method & Notes',index=False)
autofmt(out)
print("wrote",out)
print(f"\nFaculty: {len(fac)} | deposit rows: {len(det)} | concerning rows: {len(conc)} | NON-EXEMPT rows: {len(ne)} across {ne['pmid'].nunique()} pubs, {ne[['lastName','firstName']].drop_duplicates().shape[0]} faculty")
print(f"COC-coauthor: {full[full['coc_coauthor']==1]['pmid'].nunique()} pubs; COC + open-omic: {full[full['_coc_openomic']]['pmid'].nunique()} pubs; faculty on COC-coauthor pubs: {full[full['coc_coauthor']==1][['lastName','firstName']].drop_duplicates().shape[0]}")
print(f"NIH-funded depositing pubs: {full.drop_duplicates('pmid')['nih_funded'].sum()}/{full['pmid'].nunique()}")
print("\nTop 12 faculty by NON-EXEMPT open-omic deposits:")
print(fac[fac['Non-exempt open-omic']>0][['CWID','Last name','First name','Department','Non-exempt open-omic','Concerning deposits','Controlled (compliant)']].head(12).to_string(index=False))
