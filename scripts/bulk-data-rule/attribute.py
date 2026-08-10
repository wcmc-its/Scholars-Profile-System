"""People + metadata + granular sub-typing. Unifies academic + preprint deposits, attributes each to the
WCM full-time first/last author(s), joins DB metadata, fetches grants/abstract/MeSH, and writes a person- and
article-level XLSX. ponytail: pandas joins + one efetch pass; DB already has doi/journal/citations."""
import os, time, urllib.request, urllib.parse
import xml.etree.ElementTree as ET
import pandas as pd
from sqlalchemy import create_engine, text
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import date
import catalog, taxonomy

OUT=os.path.dirname(os.path.abspath(__file__)); PROJ=os.path.expanduser("~/Dropbox/Projects/Bulk Data Rule"); API=os.environ.get("PUBMED_API_KEY","")
engine=create_engine(f"mysql+pymysql://{os.environ['DB_USERNAME']}:{os.environ['DB_PASSWORD']}@{os.environ['DB_HOST']}/{os.environ['DB_NAME']}")

# ---------- 1. unified deposit set (academic + preprint; full-text high-conf + structured) ----------
# Carries accession/resource_type through (used to stop at ['pmid','repo'] — fine for the landscape
# report's per-repo counts, not enough for a per-deposit table keyed on (repo, accession)).
def load_ft(p, provenance):
    d=pd.read_csv(p)
    d=d[(d['confidence']=='high')&(d['context']!='use')]
    # kind=='acc': a real accession pattern matched. resource_bucket!='not-doi': a DOI pattern
    # matched and DataCite-typed. Everything else (bare repo-name mentions, bare-domain URL hits)
    # has no per-record locator — matched_val is just the matched text, not an id — so it can't
    # populate accession_or_doi. Still valid signal for the aggregate landscape report; not here.
    d=d[(d['kind']=='acc') | (d['resource_bucket']!='not-doi')]
    d['accession']=d['matched_val']; d['provenance']=provenance
    return d[['pmid','repo','accession','resource_type','confidence','provenance']]
def load_db(p, provenance):
    d=pd.read_csv(p); d=d[d['year']>=2020] if 'year' in d else d
    if 'tier' in d:  # extract_databanks.py's classified+tagged shape: honor the registry/use exclusion
        d=d[(d['tier']!='REGISTRY') & (d['context']!='use')]
        d['repo']=d['canonical']
    else:            # legacy shape (e.g. preprint_databank.csv, not yet wired the same way)
        d['repo']=d['databank'].apply(lambda n: catalog.classify(name=str(n))['canonical'])
    d=d[d['accession'].notna()]  # a databank name with no accession listed has no per-record locator
    d['resource_type']=None; d['confidence']=None; d['provenance']=provenance
    return d[['pmid','repo','accession','resource_type','confidence','provenance']]
parts=[load_ft(f"{OUT}/deposits_v2.csv","fulltext-scan"), load_ft(f"{OUT}/preprint_deposits_v2.csv","fulltext-scan"),
       load_db(f"{OUT}/deposits_databank.csv","databank"), load_db(f"{OUT}/preprint_databank.csv","databank")]
dep=pd.concat(parts,ignore_index=True).drop_duplicates(['pmid','repo','accession'])
dep['pmid']=dep['pmid'].astype(int)
for col in ['tier','country','access','bucket']:
    dep[col]=dep['repo'].map(lambda r: catalog.classify(name=r)[col])
pmids=sorted(dep['pmid'].unique().tolist())
print(f"{len(dep)} deposits across {len(pmids)} pubs", flush=True)

# ---------- 2. people (WCM full-time first/last authors) ----------
inlist=",".join(str(p) for p in pmids)
ppl=pd.read_sql(text(f"""SELECT a.pmid, i.cwid, i.givenName firstName, i.surname lastName,
    i.primaryAcademicDepartment dept, i.primaryAcademicDivision division, i.primaryTitle title, a.authorPosition position
    FROM analysis_summary_author a JOIN identity i ON i.cwid=a.personIdentifier
    WHERE a.pmid IN ({inlist}) AND i.fullTimeFaculty='yes' AND a.authorPosition IN ('first','last')"""), engine)
ppl['pmid']=ppl['pmid'].astype(int)

# ---------- 3. article metadata from DB ----------
art=pd.read_sql(text(f"""SELECT pmid, articleTitle title, journalTitleVerbose journal, journalImpactScore2 impact,
    articleYear year, publicationDateStandardized pubdate, doi, pmcid, publicationTypeCanonical pubtype,
    citationCountNIH citNIH, citationCountScopus citScopus FROM analysis_summary_article WHERE pmid IN ({inlist})"""), engine)
art['pmid']=art['pmid'].astype(int)

# ---------- 4. grants + abstract + raw MeSH (efetch) ----------
EFETCH="https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
def fetch(batch):
    d={"db":"pubmed","id":",".join(map(str,batch)),"retmode":"xml"}; d.update({"api_key":API} if API else {})
    for a in range(4):
        try: return urllib.request.urlopen(urllib.request.Request(EFETCH,data=urllib.parse.urlencode(d).encode()),timeout=120).read()
        except Exception:
            if a==3: raise
            time.sleep(2*(a+1))
meta=[]
for i in range(0,len(pmids),200):
    root=ET.fromstring(fetch(pmids[i:i+200]))
    for a in root.findall(".//PubmedArticle"):
        pmid=a.findtext(".//MedlineCitation/PMID")
        mesh=[(d.findtext("DescriptorName") or "") for d in a.findall(".//MeshHeading")]
        human=any(m.lower()=="humans" for m in mesh)
        agencies=sorted({(g.findtext("Agency") or "") for g in a.findall(".//GrantList/Grant")} - {""})
        nih=any("NIH" in ag or "National Institutes" in ag for ag in agencies)
        usgov=nih or any(x in " ".join(agencies) for x in ["CDC","AHRQ","VA ","Department of Defense","NSF","PHS HHS"])
        abstract=" ".join((t.text or "") for t in a.findall(".//Abstract/AbstractText"))[:500]
        meta.append((int(pmid), human, "; ".join(agencies), nih, usgov, "|".join(m for m in mesh if m), abstract))
    time.sleep(0.12 if API else 0.34)
    if (i//200)%5==0: print(f"  meta {i+200}/{len(pmids)}", flush=True)
md=pd.DataFrame(meta,columns=["pmid","human","funders","nih_funded","usgov_funded","mesh","abstract"])

# ---------- 5. granular sub-typing per pub ----------
def subtag(row):
    coarse,subs=taxonomy.tag((row['mesh'] or "").split("|"), bool(row['human']))
    return pd.Series(["|".join(sorted(coarse)), "|".join(sorted(subs))])
md[['sensitive_cats','sensitive_subtypes']]=md.apply(subtag,axis=1)

# ---------- 6. assemble per (pmid, repo, person) ----------
dm=dep.merge(md,on='pmid',how='left').merge(art,on='pmid',how='left')
dm['repo_detail']=dm.apply(lambda r: taxonomy.resolve_repo(r['repo'], set((r['sensitive_subtypes'] or "").split("|")), bool(r['human'])), axis=1)
full=ppl.merge(dm,on='pmid',how='right')   # one row per deposit x qualifying author (right: keep deposits w/o matched FT author too)

# concern flag: country-of-concern, OR foreign-hosted, OR open deposition of a sensitive data type
SENS={'genomic','omic_other','health','biometric','geolocation'}
def concern(r):
    reasons=[]
    if r['tier']=='CONCERN': reasons.append("country-of-concern host")
    if r['tier'] in ('FOREIGN_OPEN','FOREIGN_CTRL'): reasons.append("foreign-hosted")
    sens = set((r['sensitive_cats'] or "").split("|")) & SENS
    if r['tier'] in ('US_OPEN','FOREIGN_OPEN') and sens: reasons.append("open deposition of "+",".join(sorted(sens)))
    return "; ".join(reasons)
full['concern']=full.apply(concern,axis=1)
full['name']=full['firstName'].fillna('')+" "+full['lastName'].fillna('')

full.to_csv(f"{OUT}/attributed_deposits.csv",index=False)
print(f"\nrows (deposit x author): {len(full)}  distinct faculty: {full['cwid'].nunique()}")
print(f"deposits flagged concerning: {(full['concern']!='').sum()} rows")
print("\n=== top faculty by # concerning deposits ===")
cc=full[full['concern']!='']
print(cc.groupby(['name','dept'])['pmid'].nunique().sort_values(ascending=False).head(15).to_string())
print("\n=== granular sensitive sub-types (deposit rows) ===")
allsubs={}
for s in full['sensitive_subtypes'].dropna():
    for x in s.split("|"):
        if x: allsubs[x]=allsubs.get(x,0)+1
for k,v in sorted(allsubs.items(),key=lambda x:-x[1]): print(f"  {v:4d}  {k}")
print("\n=== NIH-funded share of depositing pubs ===")
pub=full.drop_duplicates('pmid'); print(f"  NIH-funded: {pub['nih_funded'].sum()}/{len(pub)}   any US-gov: {pub['usgov_funded'].sum()}")

# ---------- 7. optional: write per-deposit rows to reciterdb.dataset_deposit ----------
# ponytail: additive, env-gated — the existing CSV/xlsx pipeline (coauthor_affil.py,
# build_people.py still read attributed_deposits.csv) is untouched either way.
# deposit_year is always the citing pub's year (no repo-native deposit-date extraction yet;
# matches the SPS DatasetDeposit model's documented "else pub year" fallback).
# cwid may be null (kept — see extract_databanks/scan2: a deposit can exist with no matched
# WCM full-time first/last author; the SPS bridge dedups DatasetDeposit by (repo, accession)
# before it ever looks at cwid, so these rows still populate the deposit catalog).
if os.environ.get("WRITE_DATASET_DEPOSIT"):
    import pymysql
    def nz(v): return None if pd.isna(v) else v
    def access_model(v):  # catalog.py's `access` is a free-text note, not the table's open|controlled enum
        return "controlled" if "controlled" in str(v).lower() else "open"
    recs=[(nz(r.cwid), r.repo, r.accession, nz(r.resource_type), nz(r.bucket), access_model(r.access),
           int(r.year) if pd.notna(r.year) else None, r.provenance, nz(r.confidence),
           nz(r.position), int(r.pmid))
          for r in full.itertuples()]
    conn=pymysql.connect(host=os.environ['DB_HOST'], user=os.environ['DB_USERNAME'],
                          password=os.environ['DB_PASSWORD'], database=os.environ['DB_NAME'])
    with conn.cursor() as cur:
        cur.executemany("""INSERT IGNORE INTO dataset_deposit
            (cwid, repository, accession_or_doi, resource_type, data_type, access_model,
             deposit_year, provenance, confidence, author_position, pmid)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""", recs)
        written=cur.rowcount
    conn.commit(); conn.close()
    print(f"\nwrote {written} new rows to reciterdb.dataset_deposit ({len(recs)} candidates, dupes ignored)")
