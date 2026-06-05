#!/usr/bin/env python3
"""
Track A — Phase 0 corpus prep.

Joins the 2022 COI reference corpus into clean intermediate CSVs that the
dependency-free Node harness (analyze.mjs) consumes. Reads ONLY from the
read-only reference directory; writes ONLY to an out dir under /tmp.

The official Conflicts-Survey export (coi_03042022_1.xlsx) is CONFIDENTIAL
(names, entities, dollar ranges, family-member flags). Nothing it contains is
written inside the repo — every output lands in OUT_DIR (default /tmp/coi-phase0).
Never `git add` the out dir.

Excel corrupted ~1,195 of 5,241 cwids by coercing numeric-looking ids into
dates (e.g. 2039-03-01). A cwid string either coerces or it doesn't, so a given
faculty's rows are all-clean or all-corrupted. We keep only valid-cwid rows;
faculty whose cwid is corrupted are dropped from the study population entirely
(rather than mis-scored as "disclosed nothing").

Usage:
    python3 prep.py [REF_DIR] [OUT_DIR]
Defaults:
    REF_DIR = ~/Dropbox/Index/Conflicts of Interest : External Relationships
    OUT_DIR = /tmp/coi-phase0
"""
import csv
import io
import os
import re
import sys
from openpyxl import load_workbook

REF_DIR = sys.argv[1] if len(sys.argv) > 1 else os.path.expanduser(
    "~/Dropbox/Index/Conflicts of Interest : External Relationships"
)
OUT_DIR = sys.argv[2] if len(sys.argv) > 2 else "/tmp/coi-phase0"

KNOWNS = os.path.join(REF_DIR, "KnownsPubs2019+FullTimeFaculty.csv")
CONFLICTS_TAB = os.path.join(REF_DIR, "PubMedDownload-2021-06-27.tab")
COI_XLSX = os.path.join(REF_DIR, "coi_03042022_1.xlsx")

CWID_RE = re.compile(r"[a-z]{2,4}\d{2,4}")


def looks_cwid(v):
    return v is not None and bool(CWID_RE.fullmatch(str(v)))


def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    # ---- conflicts.csv (pmid -> statement) ----
    with open(CONFLICTS_TAB, encoding="utf-8", errors="replace") as f:
        rdr = csv.reader(io.StringIO(f.read()), delimiter="\t")
        conflict_rows = [(r[0], r[1] if len(r) > 1 else "")
                         for r in rdr if r and r[0].isdigit()]
    pmids_with_text = {p for p, t in conflict_rows if t.strip()}
    with open(os.path.join(OUT_DIR, "conflicts.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["pmid", "statement"])
        w.writerows(conflict_rows)

    # ---- disclosed set, valid cwid only ----
    wb = load_workbook(COI_XLSX, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h) for h in rows[0]]
    idx = {h: i for i, h in enumerate(hdr)}
    col = lambda r, name: (r[idx[name]] if idx.get(name) is not None and idx[name] < len(r) else None)

    disclosed = []          # one row per disclosed relationship
    fac_name = {}           # cwid -> (last, first)
    corrupted_cwid_rows = 0
    for r in rows[1:]:
        cw = r[0]
        if not looks_cwid(cw):
            corrupted_cwid_rows += 1
            continue
        cw = str(cw)
        last = str(col(r, "Last Name") or "").strip()
        first = str(col(r, "First Name") or "").strip()
        fac_name.setdefault(cw, (last, first))
        disclosed.append({
            "cwid": cw,
            "last": last,
            "first": first,
            "entity": str(col(r, "entity") or "").strip(),
            "activityType": str(col(r, "Activity Type") or "").strip(),
            "value": str(col(r, "value") or "").strip(),
            "relatesTo": str(col(r, "Activity Relates To") or "").strip(),
        })
    coi_valid_cwids = set(fac_name)

    # ---- knowns.csv (cwid,pmid,year) ----
    with open(KNOWNS, encoding="utf-8", errors="replace") as f:
        knowns = list(csv.DictReader(f))
    knowns_cwids = {r["personIdentifier"] for r in knowns}

    # ---- study population: valid cwid in BOTH files ----
    study = sorted(knowns_cwids & coi_valid_cwids)
    study_set = set(study)

    with open(os.path.join(OUT_DIR, "knowns.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["cwid", "pmid", "year"])
        for r in knowns:
            if r["personIdentifier"] in study_set:
                w.writerow([r["personIdentifier"], r["pmid"], r.get("articleYear", "")])

    with open(os.path.join(OUT_DIR, "disclosed.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["cwid", "last", "first", "entity", "activityType", "value", "relatesTo"])
        w.writeheader()
        for d in disclosed:
            if d["cwid"] in study_set:
                w.writerow(d)

    with open(os.path.join(OUT_DIR, "faculty.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["cwid", "last", "first"])
        for cw in study:
            last, first = fac_name[cw]
            w.writerow([cw, last, first])

    # ---- coverage stats ----
    study_pmids = {r["pmid"] for r in knowns if r["personIdentifier"] in study_set}
    covered = study_pmids & pmids_with_text
    print("=== Track A prep — corpus join ===")
    print(f"REF_DIR : {REF_DIR}")
    print(f"OUT_DIR : {OUT_DIR}  (confidential — never commit)")
    print(f".tab conflict rows ............ {len(conflict_rows)}  (with text: {len(pmids_with_text)})")
    print(f"coi rows: valid-cwid={len(disclosed)}  corrupted-cwid-dropped={corrupted_cwid_rows}")
    print(f"coi valid-cwid faculty ........ {len(coi_valid_cwids)}")
    print(f"KnownsPubs faculty ............ {len(knowns_cwids)}")
    print(f"STUDY POPULATION (both) ....... {len(study)} faculty")
    print(f"  their pmids ................. {len(study_pmids)}")
    print(f"  pmids with COI text ........ {len(covered)} ({100*len(covered)/max(1,len(study_pmids)):.1f}%)")
    print(f"  disclosed relationships .... {sum(1 for d in disclosed if d['cwid'] in study_set)}")
    print("wrote: knowns.csv conflicts.csv disclosed.csv faculty.csv")


if __name__ == "__main__":
    main()
